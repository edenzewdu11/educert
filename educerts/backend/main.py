from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Cookie, Response, Request, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from fastapi.templating import Jinja2Templates
from fastapi.exceptions import RequestValidationError
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import timedelta
import uuid
import csv
import io
import datetime
import hashlib
import random
import os
import tempfile
from xhtml2pdf import pisa
from io import BytesIO
import qrcode
import base64
from dotenv import load_dotenv
import fitz  # PyMuPDF

import models, schemas, crypto_utils, database, auth_utils, oa_logic
import pdf_utils
import pdf_hash_utils
import builtin_templates
from wps_ribbon_final import add_final_ribbon
# DISABLED: Remove other ribbon imports - keep only WPS ribbon
# import pdf_ribbon_utils
# import verification_metadata
# import pdf_javascript_templates
# from ribbon_styling import RibbonStyle, RibbonPosition

load_dotenv()

# Create tables
models.Base.metadata.create_all(bind=database.engine)

# Auto-seed admin user for production
def seed_admin_user():
    db = database.SessionLocal()
    try:
        admin_email = "app-reguser@mailinator.com"
        user = db.query(models.User).filter(models.User.email == admin_email).first()
        if not user:
            print("Creating admin user for production...")
            new_user = models.User(
                name="Eden",
                email=admin_email,
                password=auth_utils.get_password_hash("Password1"),
                is_admin=True
            )
            db.add(new_user)
            db.commit()
            print("Admin user created successfully.")
    except Exception as e:
        print(f"Error seeding admin user: {e}")
    finally:
        db.close()

seed_admin_user()

app = FastAPI(title="EduCerts API")
templates = Jinja2Templates(directory="templates")

FRONTEND_URL = os.getenv("FRONTEND_URL", "http://localhost:3000")
ENVIRONMENT = os.getenv("ENVIRONMENT", "development")

# CORS setup - origins come from env (FRONTEND_URL, comma-separated) plus local dev defaults
_default_origins = [
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "http://192.168.56.1:3000",
]
_env_origins = [o.strip() for o in os.getenv("FRONTEND_URL", "").split(",") if o.strip()]
allowed_origins = list(dict.fromkeys(_default_origins + _env_origins))

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request, exc):
    return JSONResponse(
        status_code=422,
        content={"detail": str(exc.errors()[0]["msg"]) if exc.errors() else "Validation error"}
    )

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc):
    import traceback
    print(f"GLOBAL ERROR: {exc}\n{traceback.format_exc()}")
    
    # Get origin from request
    origin = request.headers.get("origin", "http://localhost:3000")
    
    return JSONResponse(
        status_code=500,
        content={"detail": "Internal server error"},
        headers={
            "Access-Control-Allow-Origin": origin,
            "Access-Control-Allow-Credentials": "true",
            "Access-Control-Allow-Methods": "*",
            "Access-Control-Allow-Headers": "*",
        }
    )

def get_db():
    db = database.SessionLocal()
    try:
        yield db
    finally:
        db.close()

def normalize_column_name(header: str) -> str:
    """
    Normalizes a header name (lowercase, strip, underscores).
    Maps common aliases to student_name and course_name.
    """
    if not header:
        return ""
    import re
    h = str(header).lower().strip()
    h = h.replace(" ", "_").replace("-", "_")
    h = re.sub(r'[^a-z0-9_]', '', h)
    
    # Aliases for student_name
    if h in {"student_name", "student", "full_name", "name", "recipient", "recipient_name", "candidate_name", "student_fullname",
             "rollno", "roll_no", "enrn", "enrollment_no", "student_id", "reg_no", "registration_number", "studentname", "fullname"}:
        return "student_name"
    
    # Aliases for course_name
    if h in {"course_name", "course", "subject", "program", "training_name", "training", "module", "study_program",
             "subject_1", "subject_code", "course_code", "coursename", "trainingname"}:
        return "course_name"

    return h

def get_current_user_from_cookie(
    access_token: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
) -> Optional[models.User]:
    """Dependency to extract and validate the user from the HttpOnly cookie."""
    if not access_token:
        print("DEBUG AUTH: No access_token cookie found")
        return None
    
    payload = auth_utils.decode_access_token(access_token)
    if not payload:
        print("DEBUG AUTH: Failed to decode access token")
        return None
        
    username = payload.get("sub")
    if not username:
        print("DEBUG AUTH: No username in token payload")
        return None
        
    user = db.query(models.User).filter(models.User.name == username).first()
    if user:
        print(f"DEBUG AUTH: User authenticated: {user.name} (admin={user.is_admin})")
    else:
        print(f"DEBUG AUTH: User not found in database: {username}")
    return user

def require_user(current_user: Optional[models.User] = Depends(get_current_user_from_cookie)) -> models.User:
    """Dependency that requires an authenticated user."""
    if not current_user:
        raise HTTPException(
            status_code=401, 
            detail="Not authenticated. Please log in again."
        )
    return current_user

def require_admin(current_user: models.User = Depends(require_user)) -> models.User:
    """Dependency that requires an admin user."""
    if not current_user.is_admin:
        raise HTTPException(
            status_code=403, 
            detail=f"Admin access required. Current user '{current_user.name}' is not an admin."
        )
    return current_user

def generate_qr_base64(data: str):
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    return base64.b64encode(buffered.getvalue()).decode()

def activate_builtin_template(cert_type: str) -> dict:
    """
    Activate the built-in HTML template for a certificate type and return
    field metadata used by frontend and bulk mapping logic.
    """
    normalized_type = (cert_type or "certificate").lower().strip() or "certificate"
    html = builtin_templates.get_builtin_template_html(normalized_type)

    os.makedirs("user_templates", exist_ok=True)
    with open("user_templates/custom_certificate.html", "w", encoding="utf-8") as f:
        f.write(html)

    # Ensure legacy PDF template uploads do not override type-based templates.
    pdf_template_path = "user_templates/template.pdf"
    if os.path.exists(pdf_template_path):
        try:
            os.remove(pdf_template_path)
        except OSError:
            pass

    field_defs = builtin_templates.get_builtin_template_fields(normalized_type)
    input_fields = [fd["key"] for fd in field_defs]
    field_labels = {fd["key"]: fd["label"] for fd in field_defs}
    required_fields = [fd["key"] for fd in field_defs if fd["required"]]
    custom_fields = [k for k in input_fields if k not in ("student_name", "course_name")]

    system_fields = ["issued_at", "cert_id", "signature", "qr_code", "signer_name", "signer_role"]
    sig_fields = ["digital_signature", "stamp"]

    return {
        "normalized_type": normalized_type,
        "all_fields": input_fields + system_fields + sig_fields,
        "system_fields": system_fields,
        "signature_fields": sig_fields,
        "custom_fields": custom_fields,
        "input_fields": input_fields,
        "field_labels": field_labels,
        "required_fields": required_fields,
        "template_name": builtin_templates.get_builtin_template_label(normalized_type),
        "template_type": "html",
    }
# ─────────────────────────────────────────────────────────────────────────────
# Auth Endpoints
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/")
def read_root():
    return {"message": "EduCerts API — Secure Mode"}

@app.post("/api/signup")
def signup(user_data: schemas.UserCreate, db: Session = Depends(get_db)):
    if db.query(models.User).filter(models.User.email == user_data.email).first():
        raise HTTPException(status_code=400, detail="Email already registered")
    if db.query(models.User).filter(models.User.name == user_data.name).first():
        raise HTTPException(status_code=400, detail="Name already taken")

    hashed_password = auth_utils.get_password_hash(user_data.password)
    new_user = models.User(
        name=user_data.name,
        email=user_data.email,
        password=hashed_password,
        is_admin=False
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return {"id": new_user.id, "name": new_user.name, "email": new_user.email}

@app.post("/api/login")
def login(response: Response, form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(
        (models.User.name == form_data.username) | (models.User.email == form_data.username)
    ).first()

    if not user or not auth_utils.verify_password(form_data.password, user.password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
        )

    access_token_expires = timedelta(minutes=auth_utils.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = auth_utils.create_access_token(
        data={"sub": user.name}, expires_delta=access_token_expires
    )

    # Set token as secure HttpOnly cookie instead of returning it in body
    is_production = ENVIRONMENT == "production"
    # Cross-site cookies (frontend and backend on different domains) require
    # SameSite=None + Secure. In local dev keep Lax over http.
    cookie_samesite = "none" if is_production else "lax"
    response.set_cookie(
        key="access_token",
        value=access_token,
        httponly=True,           # Not accessible via JavaScript
        secure=is_production,    # HTTPS-only in production (required for SameSite=None)
        samesite=cookie_samesite,
        max_age=auth_utils.ACCESS_TOKEN_EXPIRE_MINUTES * 60
    )

    # Still return user info (but NOT the token)
    return {
        "user": {
            "id": user.id,
            "name": user.name,
            "email": user.email,
            "is_admin": user.is_admin
        }
    }

@app.get("/api/me")
def get_current_user_info(current_user: models.User = Depends(require_user)):
    """Returns the currently authenticated user from the HttpOnly cookie."""
    return {
        "id": current_user.id,
        "name": current_user.name,
        "email": current_user.email,
        "is_admin": current_user.is_admin
    }

@app.post("/api/logout")
def logout(response: Response):
    """Clears the auth cookie."""
    is_production = ENVIRONMENT == "production"
    response.delete_cookie(
        key="access_token",
        secure=is_production,
        samesite="none" if is_production else "lax",
    )
    return {"message": "Logged out successfully"}

# ─────────────────────────────────────────────────────────────────────────────
# Certificate Issuance (Phase 3: Document Registry)
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/issue", response_model=schemas.Certificate)
def issue_certificate(cert_data: schemas.CertificateCreate, db: Session = Depends(get_db)):
    cert_type = (cert_data.cert_type or "certificate").lower()
    activate_builtin_template(cert_type)
    cert_id = str(uuid.uuid4())

    # Build a generic document structure — no hardcoded transcript assumption
    raw_data = {
        "id": str(uuid.uuid4())[:8],
        "type": cert_type,
        "name": cert_data.course_name,
        "issuedOn": datetime.datetime.now().isoformat(),
        "recipient": {
            "name": cert_data.student_name,
            "studentId": cert_data.data_payload.get("student_id", "N/A")
        },
        # Spread all extra data_payload fields into the document (grade, department, gpa, etc.)
        **{k: v for k, v in cert_data.data_payload.items() if k not in ("student_id", "organization")}
    }

    organization = cert_data.data_payload.get("organization", "EduCerts Academy")
    issuers = [
        {
            "name": organization,
            "url": "https://educerts.io",
            "documentStore": "0x007d40224f6562461633ccfbaffd359ebb2fc9ba",
            "identityProof": {"type": "DNS-TXT", "location": "educerts.io"}
        }
    ]

    # Wrap and sign
    oa_doc = oa_logic.wrap_document(raw_data, issuers=issuers)
    # Even for a single cert, we treat it as a batch of 1
    merkle_root = oa_doc["signature"]["merkleRoot"]
    target_hash = oa_doc["signature"]["targetHash"]
    
    # Sign the merkle root
    sig = crypto_utils.sign_data(merkle_root)
    oa_doc["signature"]["signature"] = sig
    oa_doc["signature"]["publicKey"] = crypto_utils.get_public_key_pem()

    # Anchor Merkle Root to Document Registry
    batch_id = str(uuid.uuid4())
    doc_registry_entry = models.DocumentRegistry(
        id=batch_id,
        merkle_root=merkle_root,
        issuer_name="EduCerts Admin",
        organization=organization,
        cert_count=1,
    )
    db.add(doc_registry_entry)

    claim_pin = None

    # If a PDF template exists, render it immediately
    pdf_template_path = "user_templates/template.pdf"
    rendered_path = None
    if os.path.exists(pdf_template_path):
        verify_url = f"{FRONTEND_URL}/verify?id={cert_id}"
        qr_b64 = generate_qr_base64(verify_url)
        issued_at_str = datetime.datetime.now().strftime("%Y-%m-%d")
        
        # Build field values from the data_payload + core fields
        field_values = {
            "student_name": cert_data.student_name,
            "name": cert_data.student_name, # Alias
            "recipient": cert_data.student_name, # Alias
            "recipient_name": cert_data.student_name, # PDF template field
            "dept_head": cert_data.data_payload.get("dept_head", cert_data.data_payload.get("department_head", cert_data.data_payload.get("head_of_department", ""))), # PDF template field
            "course_name": cert_data.course_name,
            "course": cert_data.course_name, # Alias
            "subject": cert_data.course_name, # Alias
            "issued_at": issued_at_str,
            "date": issued_at_str, # Alias
            "cert_id": cert_id,
            "id": cert_id, # alias
            "id8": cert_id[:8], # short alias
            "signature": sig[:20] + "...",
            "qr_code": qr_b64,
            **cert_data.data_payload
        }
        
        print(f"DEBUG [issue_certificate]: field_values for PDF rendering:")
        for key, value in field_values.items():
            print(f"  {key}: '{value}'")
        print(f"DEBUG [issue_certificate]: dept_head value: '{field_values.get('dept_head', 'NOT FOUND')}'")
        
        
        os.makedirs("generated_certs", exist_ok=True)
        out_path = f"generated_certs/{cert_id}_base.pdf"
        try:
            pdf_utils.render_pdf_certificate(
                pdf_template_path, 
                field_values, 
                out_path,
                metadata={"cert_id": cert_id}
            )
            rendered_path = out_path
        except Exception as e:
            print(f"DEBUG: Single issuance PDF render failed: {e}")

    # Compute content hash for tamper detection
    content_hash = None
    if rendered_path and os.path.exists(rendered_path):
        try:
            # Compute SHA-256 hash of PDF content
            content_hash = pdf_hash_utils.compute_pdf_content_hash(rendered_path)
            
            # Embed hash in PDF metadata for offline verification
            pdf_hash_utils.embed_hash_in_pdf_metadata(
                rendered_path,
                content_hash,
                cert_id
            )
            print(f"DEBUG: Computed and embedded content hash: {content_hash[:8]}...")
        except Exception as e:
            # Log error but don't block issuance (graceful degradation)
            print(f"WARNING: Failed to compute/embed content hash: {e}")
            content_hash = None

    db_cert = models.Certificate(
        id=cert_id,
        student_name=cert_data.student_name,
        course_name=cert_data.course_name,
        cert_type=cert_type,
        data_payload=oa_doc,
        signature=sig,
        claim_pin=claim_pin,
        organization=organization,
        batch_id=batch_id,
        template_type="html",
        rendered_pdf_path=rendered_path,
        signing_status="unsigned",
        claimed=False,
        content_hash=content_hash  # Store hash for verification
    )
    db.add(db_cert)
    db.commit()
    db.refresh(db_cert)
    
    # ═══════════════════════════════════════════════════════════════════
    # ADD BASIC VERIFICATION RIBBON TO ISSUED CERTIFICATE
    # ═══════════════════════════════════════════════════════════════════
    if rendered_path and os.path.exists(rendered_path):
        try:
            # DISABLED: Remove unwanted blue rectangle ribbon from unsigned certificates
            # The WPS ribbon will be added only during the signing step
            # from pdf_ribbon_integration import safe_add_ribbon_to_certificate
            # from ribbon_styling import RibbonStylePresets
            # 
            # # Use minimal styling for newly issued (unsigned) certificates
            # minimal_styling = RibbonStylePresets.minimal()
            # minimal_styling.background_color = "#6b7280"  # Gray for unsigned
            # minimal_styling.logo_text = "EduCerts - Unsigned"
            # 
            # # Add basic ribbon to the issued certificate
            # ribbon_success, final_pdf_path = safe_add_ribbon_to_certificate(
            #     cert=db_cert,
            #     signed_pdf_path=rendered_path,
            #     db=db,
            #     styling=minimal_styling,
            #     use_full_verification=False
            # )
            
            # Use the original rendered path without extra ribbon
            final_pdf_path = rendered_path
            ribbon_success = True
            
            if ribbon_success:
                db_cert.rendered_pdf_path = final_pdf_path
                db.commit()
                print(f"✓ Successfully added basic ribbon to issued certificate {cert_id}")
            else:
                print(f"⚠️ Failed to add basic ribbon to issued certificate {cert_id}")
                
        except Exception as ribbon_error:
            print(f"WARNING: Basic ribbon integration failed for issued certificate {cert_id}: {ribbon_error}")
            # Continue with standard certificate - graceful degradation
    
    return db_cert


# ─────────────────────────────────────────────────────────────────────────────
# Claiming
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/claim")
def claim_certificate(claim_data: dict, db: Session = Depends(get_db), current_user: models.User = Depends(require_user)):
    pin = claim_data.get("pin")
    org = claim_data.get("organization")

    if not pin or len(pin) != 6:
        raise HTTPException(status_code=400, detail="Invalid PIN format. Must be 6 digits.")

    cert = db.query(models.Certificate).filter(
        models.Certificate.claim_pin == pin,
        models.Certificate.organization == org
    ).first()

    if not cert:
        raise HTTPException(status_code=404, detail="No certificate found for this PIN and Organization")
    if cert.revoked:
        raise HTTPException(status_code=400, detail="This certificate has been revoked")

    cert.claimed = True
    cert.student_id = current_user.id
    db.commit()
    db.refresh(cert)
    
    # Return the full cert so the frontend gets the true database ID
    return {
        "id": cert.id,
        "student_name": cert.student_name,
        "course_name": cert.course_name,
        "cert_type": cert.cert_type,
        "issued_at": cert.issued_at.isoformat() if cert.issued_at else None,
        "organization": cert.organization,
        "data_payload": cert.data_payload
    }

# ─────────────────────────────────────────────────────────────────────────────
# Verification (Phase 3: Check Document Registry)
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/verify")
def verify_certificate(request: schemas.VerificationRequest, db: Session = Depends(get_db)):
    oa_doc = None
    cert = None

    if request.certificate_id:
        # Try full UUID first
        cert = db.query(models.Certificate).filter(models.Certificate.id == request.certificate_id).first()
        
        # If not found, try short ID (prefix match)
        if not cert and len(request.certificate_id) == 8:
            cert = db.query(models.Certificate).filter(models.Certificate.id.like(f"{request.certificate_id}%")).first()
            
        if not cert:
            raise HTTPException(status_code=404, detail="Certificate not found")
        oa_doc = cert.data_payload
    elif request.pin:
        pin = request.pin.strip()
        if len(pin) != 6 or not pin.isdigit():
            raise HTTPException(status_code=400, detail="Invalid PIN format. Must be 6 digits.")
        query = db.query(models.Certificate).filter(models.Certificate.claim_pin == pin)
        if request.organization:
            query = query.filter(models.Certificate.organization == request.organization)
        cert = query.first()
        if not cert:
            raise HTTPException(status_code=404, detail="Certificate not found for this PIN")
        oa_doc = cert.data_payload
    elif request.data_payload:
        oa_doc = request.data_payload
        signature = oa_doc.get("signature", {}).get("signature")
        cert = db.query(models.Certificate).filter(models.Certificate.signature == signature).first()

    if not oa_doc:
        raise HTTPException(status_code=400, detail="Must provide certificate_id or data_payload")

    # 1. Integrity Check
    merkle_root = oa_doc.get("signature", {}).get("merkleRoot")
    target_hash = oa_doc.get("signature", {}).get("targetHash")
    proof = oa_doc.get("signature", {}).get("proof", [])
    signature = oa_doc.get("signature", {}).get("signature")
    salted_data = oa_doc.get("data", {})
    
    # Calculate target hash from data
    field_hashes = oa_logic.get_field_hashes(salted_data)
    calculated_target_hash = oa_logic.calculate_merkle_root(field_hashes)
    
    # Check if target hash matches calculations
    is_target_hash_valid = (calculated_target_hash == target_hash)
    
    # Verify target hash belongs to merkle root using proof
    is_merkle_root_valid = oa_logic.verify_merkle_proof(target_hash, proof, merkle_root)
    
    is_integrity_valid = is_target_hash_valid and is_merkle_root_valid

    # 2. Document Status
    is_issued = cert is not None
    is_not_revoked = cert and not cert.revoked

    # 3. Issuer Identity
    issuer_name = "Unknown"
    is_identity_valid = False
    issuers_data = oa_doc.get("data", {}).get("issuers", {}).get("value", [])
    if issuers_data and issuers_data[0].get("name") in ["EduCerts Academy", cert.organization if cert else ""]:
        issuer_name = f"{issuers_data[0].get('name')} (Verified)"
        is_identity_valid = True

    # 4. Signature Check
    is_signature_valid = crypto_utils.verify_signature(merkle_root, signature) if signature and merkle_root else False

    # ── Phase 3: Document Registry Check ──
    is_registry_valid = False
    if merkle_root:
        registry_entry = db.query(models.DocumentRegistry).filter(
            models.DocumentRegistry.merkle_root == merkle_root,
            models.DocumentRegistry.revoked == False
        ).first()
        is_registry_valid = registry_entry is not None
    print(f"DEBUG VERIFY: Registry Valid: {is_registry_valid}")

    all_valid = is_integrity_valid and is_issued and is_not_revoked and is_identity_valid and is_signature_valid and is_registry_valid

    return {
        "summary": {
            "all": all_valid,
            "documentStatus": is_issued and is_not_revoked,
            "documentIntegrity": is_integrity_valid,
            "issuerIdentity": is_identity_valid,
            "signature": is_signature_valid,
            "registryCheck": is_registry_valid
        },
        "data": [
            {
                "type": "DOCUMENT_INTEGRITY",
                "name": "OpenAttestationHash",
                "data": is_integrity_valid,
                "status": "VALID" if is_integrity_valid else "INVALID"
            },
            {
                "type": "DOCUMENT_STATUS",
                "name": "OpenAttestationIssued",
                "data": {"issued": is_issued, "revoked": not is_not_revoked},
                "status": "VALID" if is_issued and is_not_revoked else "INVALID"
            },
            {
                "type": "ISSUER_IDENTITY",
                "name": "DNS-TXT",
                "data": {"name": issuer_name, "id": "educerts.io"},
                "status": "VALID" if is_identity_valid else "INVALID"
            },
            {
                "type": "REGISTRY_CHECK",
                "name": "DocumentRegistry",
                "data": is_registry_valid,
                "status": "VALID" if is_registry_valid else "INVALID"
            },
            {
                "type": "SIGNATURE_CHECK",
                "name": "CryptoSignature",
                "data": is_signature_valid,
                "status": "VALID" if is_signature_valid else "INVALID"
            }
        ],
        "certificate": {
            "student_name": cert.student_name if cert else "Unknown",
            "course_name": cert.course_name if cert else "Unknown"
        }
    }

@app.post("/api/verify/pdf")
async def verify_pdf_certificate(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    COMPREHENSIVE PDF VERIFICATION using hash-based certificate matching.
    This system compares uploaded PDFs against ALL certificates in the database
    using multiple hash algorithms to detect ANY type of tampering.
    """
    
    content = await file.read()

    # Validate it's actually a PDF
    if not content.startswith(b"%PDF"):
        raise HTTPException(status_code=400, detail="Uploaded file is not a valid PDF.")

    # Save to temporary file for hash computation
    temp_fd, temp_path = tempfile.mkstemp(suffix=".pdf")
    try:
        with os.fdopen(temp_fd, 'wb') as tmp_file:
            tmp_file.write(content)
        
        print(f"🔍 COMPREHENSIVE HASH VERIFICATION for uploaded file: {file.filename}")
        
        # Step 1: Compute comprehensive hashes for uploaded file
        try:
            doc = fitz.open(temp_path)
            
            # Extract text content
            uploaded_text = ""
            for page in doc:
                uploaded_text += page.get_text()
            
            # Compute multiple hash types for maximum security
            uploaded_text_hash = hashlib.sha256(uploaded_text.encode('utf-8')).hexdigest()
            uploaded_normalized_hash = hashlib.sha256(pdf_hash_utils.normalize_pdf_text(uploaded_text).encode('utf-8')).hexdigest()
            
            with open(temp_path, 'rb') as f:
                uploaded_binary_hash = hashlib.sha256(f.read()).hexdigest()
            
            doc.close()
            
            print(f"📊 Uploaded file hashes:")
            print(f"   Text hash: {uploaded_text_hash[:16]}...")
            print(f"   Normalized hash: {uploaded_normalized_hash[:16]}...")
            print(f"   Binary hash: {uploaded_binary_hash[:16]}...")
            
        except Exception as e:
            print(f"❌ ERROR computing uploaded file hashes: {e}")
            raise HTTPException(status_code=400, detail=f"Could not process PDF: {e}")
        
        # Step 2: Check against ALL certificates in database
        print("🔍 Checking against certificate database...")
        
        all_certificates = db.query(models.Certificate).all()
        hash_match_found = False
        matched_cert = None
        match_details = {}
        
        for cert in all_certificates:
            if not cert.rendered_pdf_path or not os.path.exists(cert.rendered_pdf_path):
                continue
            
            try:
                # Compute hashes for database certificate
                cert_doc = fitz.open(cert.rendered_pdf_path)
                cert_text = ""
                for page in cert_doc:
                    cert_text += page.get_text()
                
                cert_text_hash = hashlib.sha256(cert_text.encode('utf-8')).hexdigest()
                cert_normalized_hash = hashlib.sha256(pdf_hash_utils.normalize_pdf_text(cert_text).encode('utf-8')).hexdigest()
                
                with open(cert.rendered_pdf_path, 'rb') as f:
                    cert_binary_hash = hashlib.sha256(f.read()).hexdigest()
                
                cert_doc.close()
                
                # Check for matches
                text_match = uploaded_text_hash == cert_text_hash
                normalized_match = uploaded_normalized_hash == cert_normalized_hash
                binary_match = uploaded_binary_hash == cert_binary_hash
                
                if text_match or normalized_match or binary_match:
                    print(f"✅ HASH MATCH FOUND with certificate: {cert.id}")
                    print(f"   Student: {cert.student_name}")
                    print(f"   Course: {cert.course_name}")
                    print(f"   Text match: {text_match}")
                    print(f"   Normalized match: {normalized_match}")
                    print(f"   Binary match: {binary_match}")
                    
                    hash_match_found = True
                    matched_cert = cert
                    match_details = {
                        'text_match': text_match,
                        'normalized_match': normalized_match,
                        'binary_match': binary_match,
                        'cert_text_hash': cert_text_hash,
                        'cert_normalized_hash': cert_normalized_hash,
                        'cert_binary_hash': cert_binary_hash
                    }
                    break
                    
            except Exception as e:
                print(f"⚠️  Error checking certificate {cert.id}: {e}")
                continue
        
        if not hash_match_found:
            print("❌ NO HASH MATCHES FOUND - Certificate is TAMPERED, FAKE, or NOT FROM THIS SYSTEM")
            return {
                "summary": {
                    "all": False,
                    "documentStatus": False,
                    "documentIntegrity": False,
                    "issuerIdentity": False,
                    "signature": False,
                    "registryCheck": False,
                    "contentIntegrity": False
                },
                "data": [
                    {
                        "type": "CONTENT_INTEGRITY",
                        "name": "ComprehensiveHashVerification",
                        "data": {
                            "uploaded_text_hash": uploaded_text_hash,
                            "uploaded_normalized_hash": uploaded_normalized_hash,
                            "uploaded_binary_hash": uploaded_binary_hash,
                            "database_match_found": False,
                            "certificates_checked": len(all_certificates),
                            "reason": "No matching certificate found in database - document is tampered, fake, or not issued by this system"
                        },
                        "status": "INVALID"
                    },
                    {
                        "type": "DOCUMENT_INTEGRITY",
                        "name": "OpenAttestationHash",
                        "data": False,
                        "status": "INVALID"
                    },
                    {
                        "type": "DOCUMENT_STATUS",
                        "name": "OpenAttestationIssued",
                        "data": {"issued": False, "revoked": True},
                        "status": "INVALID"
                    },
                    {
                        "type": "ISSUER_IDENTITY",
                        "name": "DNS-TXT",
                        "data": {"name": "Unknown - Invalid Certificate", "id": "unknown"},
                        "status": "INVALID"
                    },
                    {
                        "type": "REGISTRY_CHECK",
                        "name": "DocumentRegistry",
                        "data": False,
                        "status": "INVALID"
                    },
                    {
                        "type": "SIGNATURE_CHECK",
                        "name": "CryptoSignature",
                        "data": False,
                        "status": "INVALID"
                    }
                ],
                "certificate": {
                    "student_name": "INVALID - Tampered/Fake Certificate",
                    "course_name": "INVALID - Tampered/Fake Certificate"
                }
            }
        
        # Step 3: Hash match found - certificate is VALID
        print(f"✅ Certificate VERIFIED - Hash match found with cert: {matched_cert.id}")
        
        # Run standard verification checks on the matched certificate
        request = schemas.VerificationRequest(certificate_id=matched_cert.id)
        verification_result = verify_certificate(request, db)
        
        # Override content integrity to VALID since hash match was found
        verification_result["summary"]["contentIntegrity"] = True
        verification_result["summary"]["all"] = True  # Force to valid since hash matched
        
        # Add comprehensive hash verification details
        verification_result["data"].insert(0, {
            "type": "CONTENT_INTEGRITY",
            "name": "ComprehensiveHashVerification",
            "data": {
                "database_match_found": True,
                "matched_certificate": matched_cert.id,
                "match_types": {
                    "text_match": match_details['text_match'],
                    "normalized_match": match_details['normalized_match'],
                    "binary_match": match_details['binary_match']
                },
                "uploaded_hashes": {
                    "text_hash": uploaded_text_hash,
                    "normalized_hash": uploaded_normalized_hash,
                    "binary_hash": uploaded_binary_hash
                },
                "certificate_hashes": {
                    "text_hash": match_details['cert_text_hash'],
                    "normalized_hash": match_details['cert_normalized_hash'],
                    "binary_hash": match_details['cert_binary_hash']
                },
                "certificates_checked": len(all_certificates)
            },
            "status": "VALID"
        })
        
        # Update certificate info
        verification_result["certificate"] = {
            "student_name": matched_cert.student_name,
            "course_name": matched_cert.course_name
        }
        
        return verification_result
        
    finally:
        # Clean up temporary file
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass

# ─────────────────────────────────────────────────────────────────────────────
# Certificate CRUD
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/certificates", response_model=List[schemas.Certificate])
def get_all_certificates(db: Session = Depends(get_db)):
    return db.query(models.Certificate).order_by(models.Certificate.issued_at.desc()).all()

@app.get("/api/certificate/{cert_id}", response_model=schemas.Certificate)
def get_certificate_by_id(cert_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(require_user)):
    """Returns a single certificate by its UUID. Used for deep-linking into the signing step."""
    cert = db.query(models.Certificate).filter(models.Certificate.id == cert_id).first()
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")
    return cert

@app.get("/api/certificates/{student_name}", response_model=List[schemas.Certificate])
def get_student_certificates(student_name: str, db: Session = Depends(get_db), current_user: models.User = Depends(require_user)):
    """
    Returns certificates for a student. Checks by name or by student_id.
    """
    # Try by student_id first for reliability
    certs = db.query(models.Certificate).filter(
        models.Certificate.student_id == current_user.id
    ).all()
    
    if not certs:
        # Fallback to name search
        certs = db.query(models.Certificate).filter(
            models.Certificate.student_name == student_name
        ).all()
        
    return certs

@app.post("/api/revoke/{cert_id}")
def revoke_certificate(cert_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(require_admin)):
    cert = db.query(models.Certificate).filter(models.Certificate.id == cert_id).first()
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")
    cert.revoked = True
    # Also revoke the batch in Document Registry
    if cert.batch_id:
        registry = db.query(models.DocumentRegistry).filter(models.DocumentRegistry.id == cert.batch_id).first()
        if registry:
            registry.revoked = True
    db.commit()
    return {"message": "Certificate revoked and removed from Document Registry"}

@app.delete("/api/certificates/{cert_id}")
def delete_certificate(cert_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(require_admin)):
    cert = db.query(models.Certificate).filter(models.Certificate.id == cert_id).first()
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")
    
    # Optional: Delete the physical file if it exists
    if cert.rendered_pdf_path and os.path.exists(cert.rendered_pdf_path):
        try:
            os.remove(cert.rendered_pdf_path)
        except:
            pass
            
    db.delete(cert)
    db.commit()
    return {"message": "Certificate deleted successfully"}

@app.post("/api/certificates/bulk-revoke")
def bulk_revoke_certificates(request: schemas.BulkActionRequest, db: Session = Depends(get_db), current_user: models.User = Depends(require_admin)):
    certs = db.query(models.Certificate).filter(models.Certificate.id.in_(request.cert_ids)).all()
    for cert in certs:
        cert.revoked = True
        if cert.batch_id:
            registry = db.query(models.DocumentRegistry).filter(models.DocumentRegistry.id == cert.batch_id).first()
            if registry:
                registry.revoked = True
    db.commit()
    return {"message": f"Successfully revoked {len(certs)} certificates"}

@app.post("/api/certificates/bulk-delete")
def bulk_delete_certificates(request: schemas.BulkActionRequest, db: Session = Depends(get_db), current_user: models.User = Depends(require_admin)):
    print(f"DEBUG BULK DELETE: User {current_user.name} (admin: {current_user.is_admin}) attempting to delete {len(request.cert_ids)} certificates")
    print(f"DEBUG BULK DELETE: Certificate IDs: {request.cert_ids}")
    
    certs = db.query(models.Certificate).filter(models.Certificate.id.in_(request.cert_ids)).all()
    count = 0
    for cert in certs:
        if cert.rendered_pdf_path and os.path.exists(cert.rendered_pdf_path):
            try:
                os.remove(cert.rendered_pdf_path)
            except:
                pass
        db.delete(cert)
        count += 1
    db.commit()
    
    result = {"message": f"Successfully deleted {count} document(s)"}
    print(f"DEBUG BULK DELETE: Result: {result}")
    return result

# ─────────────────────────────────────────────────────────────────────────────
# Document Registry API
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/registry")
def get_document_registry(db: Session = Depends(get_db)):
    """Returns all anchored Merkle Roots — simulates querying the Document Store smart contract."""
    entries = db.query(models.DocumentRegistry).order_by(models.DocumentRegistry.anchored_at.desc()).all()
    return [
        {
            "id": e.id,
            "merkle_root": e.merkle_root,
            "issuer": e.issuer_name,
            "organization": e.organization,
            "cert_count": e.cert_count,
            "anchored_at": e.anchored_at.isoformat() if e.anchored_at else None,
            "revoked": e.revoked
        }
        for e in entries
    ]

# ─────────────────────────────────────────────────────────────────────────────
# Misc
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/apply/challenge")
def get_apply_challenge():
    return {
        "challenge": str(uuid.uuid4()),
        "required_fields": ["student_name", "course_name"],
        "message": "Scan with your Wallet App to verify your Degree."
    }

@app.post("/api/templates/upload")
async def upload_template(file: UploadFile = File(...)):
    raise HTTPException(
        status_code=410,
        detail="Template upload is disabled. Select a certificate type and the built-in template will be applied automatically.",
    )


# ─────────────────────────────────────────────────────────────────────────────
# PDF Template Upload & Parsing
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/templates/upload-pdf")
async def upload_pdf_template(file: UploadFile = File(...)):
    raise HTTPException(
        status_code=410,
        detail="PDF template upload is disabled. Select a certificate type and the built-in template will be applied automatically.",
    )

@app.post("/api/templates/select")
async def select_builtin_template(cert_type: str = Form("certificate")):
    """
    Activate a built-in certificate template for the given category.

    Instead of uploading a template, the admin selects a certificate category
    and the matching professionally designed template is activated. Returns the
    same field metadata shape as /api/templates/parse so the frontend flow is
    unchanged.
    """
    meta = activate_builtin_template(cert_type)
    return {
        "all_fields": meta["all_fields"],
        "system_fields": meta["system_fields"],
        "signature_fields": meta["signature_fields"],
        "custom_fields": meta["custom_fields"],
        "input_fields": meta["input_fields"],
        "field_labels": meta["field_labels"],
        "required_fields": meta["required_fields"],
        "template_name": meta["template_name"],
        "template_type": meta["template_type"],
    }


@app.post("/api/templates/parse")
async def parse_template(file: UploadFile = File(...)):
    raise HTTPException(
        status_code=410,
        detail="Template parsing from upload is disabled. Use /api/templates/select to load built-in templates by certificate type.",
    )


@app.post("/api/templates/bulk-issue")
async def bulk_issue_from_template(
    file: UploadFile = File(...),
    cert_type: str = Form("certificate"),
    db: Session = Depends(get_db)
):
    """
    Reads a CSV file and issues one certificate per row.
    The built-in template for the selected cert_type is used automatically.
    CSV column names are mapped directly to that type's field set.
    """

    if not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only CSV files are allowed")

    cert_type = (cert_type or "certificate").lower().strip() or "certificate"

    template_meta = activate_builtin_template(cert_type)
    template_fields = set(template_meta["input_fields"])
    use_pdf = False

    # Parse the file
    content_bytes = await file.read()
    content = content_bytes.decode("utf-8", errors="ignore")
    csv_reader = csv.DictReader(io.StringIO(content))
    raw_rows = list(csv_reader)

    if not raw_rows:
        raise HTTPException(status_code=400, detail="CSV file is empty")

    rows = raw_rows
    headers = list(rows[0].keys())
    name_col = next((h for h in headers if normalize_column_name(h) == "student_name"), None)
    if not name_col:
        name_col = next((h for h in headers if "name" in h.lower() or "roll" in h.lower() or "id" in h.lower()), None)
    
    course_col = next((h for h in headers if normalize_column_name(h) == "course_name"), None)
    if not course_col:
        course_col = next((h for h in headers if "course" in h.lower() or "subject" in h.lower() or "prog" in h.lower()), None)

    # Missing logic: Define needed variables
    batch_id = str(uuid.uuid4())
    organization = rows[0].get("organization", "EduCerts Academy").strip() or "EduCerts Academy"
    issued_certs = []
    system_auto = {"issued_at", "cert_id", "signature", "qr_code", "digital_signature", "stamp"}
    os.makedirs("generated_certs", exist_ok=True)

    # 1. First Pass: Create all OA documents and collect target hashes
    batch_data = []
    target_hashes = []
    
    total_rows = len(rows)
    for idx, row in enumerate(rows):
        print(f"DEBUG: Processing CSV row {idx+1}/{total_rows}...")
        student_name = row.get(name_col, "").strip() if name_col else "Student"
        course_name = row.get(course_col, "").strip() if course_col else "Course"
        
        data_payload_fields = {}
        row_keys_normalized = {normalize_column_name(k): k for k in row.keys()}
        
        for field in template_fields:
            if field in system_auto: continue
            
            # 1. Try exact match
            if field in row:
                data_payload_fields[field] = str(row[field]).strip() if row[field] is not None else ""
                continue
                
            # 2. Try normalized match
            f_norm = normalize_column_name(field)
            if f_norm in row_keys_normalized:
                val = row[row_keys_normalized[f_norm]]
                data_payload_fields[field] = str(val).strip() if val is not None else ""
                continue
            
            # 3. Handle common aliases explicitly
            if f_norm == "student_name" and name_col:
                data_payload_fields[field] = str(row[name_col]).strip() if row[name_col] is not None else ""
            elif f_norm == "course_name" and course_col:
                data_payload_fields[field] = str(row[course_col]).strip() if row[course_col] is not None else ""

        curr_cert_type = cert_type
        curr_organization = row.get("organization", organization).strip() or organization

        cert_id = str(uuid.uuid4())
        raw_data = {
            "id": cert_id[:12],
            "type": curr_cert_type,
            "name": course_name,
            "issuedOn": datetime.datetime.now().isoformat(),
            "recipient": {"name": student_name, "studentId": row.get("student_id", "N/A")},
            **{k: v for k, v in data_payload_fields.items() if k not in ("student_id", "organization")}
        }
        issuers = [{"name": curr_organization, "url": "https://educerts.io",
                    "documentStore": "0x007d40224f6562461633ccfbaffd359ebb2fc9ba",
                    "identityProof": {"type": "DNS-TXT", "location": "educerts.io"}}]

        oa_doc = oa_logic.wrap_document(raw_data, issuers=issuers)
        target_hashes.append(oa_doc["signature"]["targetHash"])
        
        batch_data.append({
            "cert_id": cert_id,
            "student_name": student_name,
            "course_name": course_name,
            "curr_cert_type": curr_cert_type,
            "curr_organization": curr_organization,
            "oa_doc": oa_doc,
            "data_payload_fields": data_payload_fields
        })

    # 2. Batch Cryptography
    batch_merkle_root = oa_logic.calculate_merkle_root(target_hashes)
    batch_sig = crypto_utils.sign_data(batch_merkle_root)
    public_key_pem = crypto_utils.get_public_key_pem()

    # Anchor Batch to Document Registry
    db.add(models.DocumentRegistry(
        id=batch_id, 
        merkle_root=batch_merkle_root,
        issuer_name="EduCerts Admin", 
        organization=organization, 
        cert_count=len(batch_data)
    ))

    # 3. Second Pass: Generate Proofs, Render PDFs and Save to DB
    for idx, item in enumerate(batch_data):
        oa_doc = item["oa_doc"]
        target_hash = oa_doc["signature"]["targetHash"]
        
        proof = oa_logic.get_merkle_proof(target_hashes, target_hash)
        
        oa_doc["signature"]["merkleRoot"] = batch_merkle_root
        oa_doc["signature"]["proof"] = proof
        oa_doc["signature"]["signature"] = batch_sig
        oa_doc["signature"]["publicKey"] = public_key_pem

        # Render PDF
        rendered_path = None
        if use_pdf:
            issued_at_str = datetime.datetime.now().strftime("%Y-%m-%d")
            field_values = {
                "student_name": item["student_name"],
                "name": item["student_name"], # Alias
                "recipient": item["student_name"], # Alias
                "recipient_name": item["student_name"], # PDF template field
                "dept_head": item["data_payload_fields"].get("dept_head", ""), # PDF template field
                "course_name": item["course_name"],
                "course": item["course_name"], # Alias
                "subject": item["course_name"], # Alias
                "issued_at": issued_at_str,
                "date": issued_at_str, # Alias
                "cert_id": item["cert_id"],
                "id": item["cert_id"],
                "id8": item["cert_id"][:8],
                "signature": batch_sig[:20] + "...",
                **item["data_payload_fields"]
            }
            out_path = f"generated_certs/{item['cert_id']}_base.pdf"
            try:
                pdf_utils.render_pdf_certificate(
                    template_path, 
                    field_values, 
                    out_path, 
                    placeholder_map=placeholder_map,
                    metadata={"cert_id": item["cert_id"]}
                )
                rendered_path = out_path
            except Exception as e:
                print(f"DEBUG: PDF RENDER ERROR: {e}")
                rendered_path = None

        db_cert = models.Certificate(
            id=item["cert_id"], 
            student_name=item["student_name"], 
            course_name=item["course_name"],
            cert_type=item["curr_cert_type"], 
            data_payload=oa_doc, 
            signature=batch_sig,
            claim_pin="".join([str(random.randint(0, 9)) for _ in range(6)]), 
            organization=item["curr_organization"], 
            batch_id=batch_id,
            template_type="html",
            rendered_pdf_path=rendered_path,
            signing_status="unsigned"
        )
        db.add(db_cert)
        issued_certs.append({
            "id": item["cert_id"], 
            "student_name": item["student_name"],
            "course_name": item["course_name"], 
            "signing_status": "unsigned"
        })
        
        if (idx + 1) % 10 == 0:
            db.commit()

    db.commit()
    return {
        "message": f"{len(issued_certs)} certificates issued from template",
        "count": len(issued_certs),
        "certificates": issued_certs
    }


@app.post("/api/templates/bulk-issue-excel")
async def bulk_issue_from_excel(
    file: UploadFile = File(...),
    cert_type: str = Form("certificate"),
    db: Session = Depends(get_db)
):
    """
    Reads an Excel (.xlsx) OR CSV file and issues one certificate per row.
    Uses the built-in template matching the selected cert_type.
    """
    filename_lower = file.filename.lower()
    if not (filename_lower.endswith(".xlsx") or filename_lower.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .csv files are allowed")

    cert_type = (cert_type or "certificate").lower().strip() or "certificate"

    template_meta = activate_builtin_template(cert_type)
    template_fields = set(template_meta["input_fields"])
    use_pdf = False

    # Parse the file
    content_bytes = await file.read()
    raw_rows = []
    if filename_lower.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(content_bytes))
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
    else:
        content_str = content_bytes.decode("utf-8", errors="ignore")
        csv_reader = csv.DictReader(io.StringIO(content_str))
        raw_rows = list(csv_reader)

    if not raw_rows:
        raise HTTPException(status_code=400, detail="File is empty")
    
    rows = raw_rows
    headers = list(rows[0].keys())
    name_col = next((h for h in headers if normalize_column_name(h) == "student_name"), None)
    if not name_col:
        name_col = next((h for h in headers if "name" in h.lower() or "roll" in h.lower() or "id" in h.lower()), None)
    
    course_col = next((h for h in headers if normalize_column_name(h) == "course_name"), None)
    if not course_col:
        course_col = next((h for h in headers if "course" in h.lower() or "subject" in h.lower() or "prog" in h.lower() or "cent" in h.lower()), None)

    # 1 batch for the whole bulk upload
    batch_id = str(uuid.uuid4())
    organization = rows[0].get("organization", "EduCerts Academy").strip() or "EduCerts Academy"

    issued_certs = []
    system_auto = {"issued_at", "cert_id", "signature", "qr_code", "digital_signature", "stamp"}
    os.makedirs("generated_certs", exist_ok=True)

    # 1. First Pass: Create all OA documents and collect target hashes
    batch_data = []
    target_hashes = []
    
    total_rows = len(rows)
    for idx, row in enumerate(rows):
        print(f"DEBUG: Processing Excel row {idx+1}/{total_rows}...")
        student_name = row.get(name_col, "").strip() if name_col else "Student"
        course_name = row.get(course_col, "").strip() if course_col else "Course"
        
        data_payload_fields = {}
        row_keys_normalized = {normalize_column_name(k): k for k in row.keys()}
        
        for field in template_fields:
            if field in system_auto: continue
            
            # 1. Try exact match
            if field in row:
                data_payload_fields[field] = str(row[field]).strip() if row[field] is not None else ""
                continue
                
            # 2. Try normalized match
            f_norm = normalize_column_name(field)
            if f_norm in row_keys_normalized:
                col_name = row_keys_normalized[f_norm]
                val = row[col_name]
                data_payload_fields[field] = str(val).strip() if val is not None else ""
                print(f"  - Mapped '{field}' (norm: '{f_norm}') from column '{col_name}'")
                continue
            
            # 3. Handle common aliases explicitly
            if f_norm == "student_name" and name_col:
                data_payload_fields[field] = str(row[name_col]).strip() if row[name_col] is not None else ""
                print(f"  - Mapped '{field}' from name_col '{name_col}'")
            elif f_norm == "course_name" and course_col:
                data_payload_fields[field] = str(row[course_col]).strip() if row[course_col] is not None else ""
                print(f"  - Mapped '{field}' from course_col '{course_col}'")
            else:
                print(f"  - WARNING: Field '{field}' NOT mapped")

        curr_cert_type = cert_type
        curr_organization = row.get("organization", organization).strip() or organization

        cert_id = str(uuid.uuid4())
        raw_data = {
            "id": cert_id[:12],
            "type": curr_cert_type,
            "name": course_name,
            "issuedOn": datetime.datetime.now().isoformat(),
            "recipient": {"name": student_name, "studentId": row.get("student_id", "N/A")},
            **{k: v for k, v in data_payload_fields.items() if k not in ("student_id", "organization")}
        }
        issuers = [{"name": curr_organization, "url": "https://educerts.io",
                    "documentStore": "0x007d40224f6562461633ccfbaffd359ebb2fc9ba",
                    "identityProof": {"type": "DNS-TXT", "location": "educerts.io"}}]

        oa_doc = oa_logic.wrap_document(raw_data, issuers=issuers)
        target_hashes.append(oa_doc["signature"]["targetHash"])
        
        batch_data.append({
            "cert_id": cert_id,
            "student_name": student_name,
            "course_name": course_name,
            "curr_cert_type": curr_cert_type,
            "curr_organization": curr_organization,
            "oa_doc": oa_doc,
            "data_payload_fields": data_payload_fields
        })

    # 2. Batch Cryptography
    batch_merkle_root = oa_logic.calculate_merkle_root(target_hashes)
    batch_sig = crypto_utils.sign_data(batch_merkle_root)
    public_key_pem = crypto_utils.get_public_key_pem()

    # Anchor Batch to Document Registry
    db.add(models.DocumentRegistry(
        id=batch_id, 
        merkle_root=batch_merkle_root,
        issuer_name="EduCerts Admin", 
        organization=organization, 
        cert_count=len(batch_data)
    ))

    # 3. Second Pass: Generate Proofs, Render PDFs and Save to DB
    for idx, item in enumerate(batch_data):
        oa_doc = item["oa_doc"]
        target_hash = oa_doc["signature"]["targetHash"]
        
        # Calculate individual Merkle Proof path
        proof = oa_logic.get_merkle_proof(target_hashes, target_hash)
        
        # Update OA document with batch info
        oa_doc["signature"]["merkleRoot"] = batch_merkle_root
        oa_doc["signature"]["proof"] = proof
        oa_doc["signature"]["signature"] = batch_sig
        oa_doc["signature"]["publicKey"] = public_key_pem

        # Render PDF
        rendered_path = None
        if use_pdf:
            issued_at_str = datetime.datetime.now().strftime("%Y-%m-%d")
            field_values = {
                "student_name": item["student_name"],
                "name": item["student_name"], # Alias
                "recipient": item["student_name"], # Alias
                "recipient_name": item["student_name"], # PDF template field
                "dept_head": item["data_payload_fields"].get("dept_head", ""), # PDF template field
                "course_name": item["course_name"],
                "course": item["course_name"], # Alias
                "subject": item["course_name"], # Alias
                "issued_at": issued_at_str,
                "date": issued_at_str, # Alias
                "cert_id": item["cert_id"],
                "id": item["cert_id"],
                "id8": item["cert_id"][:8],
                "signature": batch_sig[:20] + "...",
                **item["data_payload_fields"]
            }
            out_path = f"generated_certs/{item['cert_id']}_base.pdf"
            try:
                print(f"DEBUG: Rendering PDF for cert {item['cert_id']}...")
                pdf_utils.render_pdf_certificate(
                    template_path, 
                    field_values, 
                    out_path, 
                    placeholder_map=placeholder_map,
                    metadata={"cert_id": item["cert_id"]}
                )
                rendered_path = out_path
                print(f"DEBUG: PDF rendered successfully: {out_path}")
            except Exception as e:
                import traceback
                print(f"DEBUG: PDF RENDER ERROR: {e}")
                traceback.print_exc()
                rendered_path = None

        db_cert = models.Certificate(
            id=item["cert_id"], 
            student_name=item["student_name"], 
            course_name=item["course_name"],
            cert_type=item["curr_cert_type"], 
            data_payload=oa_doc, 
            signature=batch_sig,
            claim_pin=None,
            organization=item["curr_organization"], 
            batch_id=batch_id,
            template_type="html",
            rendered_pdf_path=rendered_path,
            signing_status="unsigned"
        )
        db.add(db_cert)
        issued_certs.append({
            "id": item["cert_id"], 
            "student_name": item["student_name"],
            "course_name": item["course_name"], 
            "signing_status": "unsigned"
        })
        
        if (idx + 1) % 10 == 0:
            print(f"DEBUG: Committing progress at row {idx+1}...")
            db.commit()

    db.commit()
    print(f"DEBUG: Finished bulk issuance. {len(issued_certs)} certs created.")
    return {
        "message": f"{len(issued_certs)} certificates issued",
        "count": len(issued_certs),
        "certificates": issued_certs
    }


# ─────────────────────────────────────────────────────────────────────────────
# Digital Signature Endpoints
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/sign/upload")
async def upload_signature_assets(
    signature_file: Optional[UploadFile] = File(None),
    stamp_file: Optional[UploadFile] = File(None),
    signer_name: str = Form(...),
    signer_role: str = Form(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin)
):
    """
    Upload a digital signature image and/or stamp image.
    Stores them under user_templates/ and saves a DigitalSignatureRecord.
    """
    os.makedirs("user_templates", exist_ok=True)
    sig_path = None
    stamp_path = None

    if signature_file and signature_file.filename:
        sig_bytes = await signature_file.read()
        sig_path = f"user_templates/signature_{current_user.id}.png"
        with open(sig_path, "wb") as f:
            f.write(sig_bytes)

    if stamp_file and stamp_file.filename:
        stamp_bytes = await stamp_file.read()
        stamp_path = f"user_templates/stamp_{current_user.id}.png"
        with open(stamp_path, "wb") as f:
            f.write(stamp_bytes)

    record = models.DigitalSignatureRecord(
        signer_name=signer_name,
        signer_role=signer_role,
        signature_path=sig_path,
        stamp_path=stamp_path
    )
    db.add(record)
    db.commit()
    db.refresh(record)

    return {
        "id": record.id,
        "signer_name": signer_name,
        "signer_role": signer_role,
        "has_signature": sig_path is not None,
        "has_stamp": stamp_path is not None,
        "uploaded_at": record.uploaded_at.isoformat() if record.uploaded_at else None
    }


@app.post("/api/sign/apply")
async def apply_digital_signatures(
    body: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin)
):
    """
    Apply digital signature/stamp to one or multiple certificates.
    Body: {
      cert_ids: [...],           # list of cert IDs to sign
      signer_name: str,
      signer_role: str,
      signature_record_id: int   # optional — use a previously uploaded record
    }
    OR if no record_id, looks for the current user's latest uploaded sig.
    """
    cert_ids = body.get("cert_ids", [])
    signer_name = body.get("signer_name", current_user.name)
    signer_role = body.get("signer_role", "Authorized Signatory")
    record_id = body.get("signature_record_id")

    # Find signature record
    if record_id:
        sig_record = db.query(models.DigitalSignatureRecord).filter(
            models.DigitalSignatureRecord.id == record_id
        ).first()
    else:
        sig_record = db.query(models.DigitalSignatureRecord).order_by(
            models.DigitalSignatureRecord.uploaded_at.desc()
        ).first()

    sig_path = sig_record.signature_path if sig_record else None
    stamp_path = sig_record.stamp_path if sig_record else None

    pdf_template_path = "user_templates/template.pdf"
    has_pdf_template = os.path.exists(pdf_template_path)

    os.makedirs("generated_certs", exist_ok=True)
    signed_certs = []
    now_iso = datetime.datetime.now().isoformat()

    for cert_id in cert_ids:
        cert = db.query(models.Certificate).filter(models.Certificate.id == cert_id).first()
        if not cert:
            continue

        signed_pdf_path = f"generated_certs/{cert_id}_signed.pdf"

        if cert.template_type == "pdf" and has_pdf_template:
            # Use PDF template for signing - FIRST re-render with all data, then apply signatures
            print(f"DEBUG: Signing PDF certificate {cert_id} - re-rendering with full data first")
            
            # Step 1: Re-render the PDF with all certificate data to ensure all fields are populated
            verify_url = f"{FRONTEND_URL}/verify?id={cert.id}"
            qr_b64 = generate_qr_base64(verify_url)
            issued_at_str = cert.issued_at.strftime("%Y-%m-%d") if cert.issued_at else datetime.datetime.now().strftime("%Y-%m-%d")
            
            # Build complete field values including dept_head from data_payload
            field_values = {
                "student_name": cert.student_name,
                "name": cert.student_name,
                "recipient": cert.student_name,
                "recipient_name": cert.student_name, # PDF template field
                "course_name": cert.course_name,
                "course": cert.course_name,
                "subject": cert.course_name,
                "issued_at": issued_at_str,
                "date": issued_at_str,
                "cert_id": cert.id,
                "id": cert.id,
                "id8": cert.id[:8],
                "signature": cert.signature[:20] + "..." if cert.signature else "N/A",
                "qr_code": qr_b64,
            }
            
            # Extract additional fields from certificate's data_payload (OpenAttestation structure)
            if cert.data_payload and isinstance(cert.data_payload, dict):
                # Handle OpenAttestation document structure
                oa_data = cert.data_payload.get("data", {})
                if isinstance(oa_data, dict):
                    # Extract values from OA salt/value structure
                    for key, salt_value_obj in oa_data.items():
                        if isinstance(salt_value_obj, dict) and "value" in salt_value_obj:
                            # Skip system fields and nested objects
                            if key not in ["id", "type", "name", "issuedOn", "recipient.name", "recipient.studentId", "issuers"]:
                                field_values[key] = str(salt_value_obj["value"]) if salt_value_obj["value"] is not None else ""
                    
                    # Specifically ensure dept_head is mapped from OA structure
                    dept_head_obj = (oa_data.get("dept_head") or 
                                   oa_data.get("department_head") or 
                                   oa_data.get("head_of_department") or 
                                   oa_data.get("hod"))
                    if dept_head_obj and isinstance(dept_head_obj, dict) and "value" in dept_head_obj:
                        field_values["dept_head"] = str(dept_head_obj["value"])
                        print(f"DEBUG: Extracted dept_head: {field_values['dept_head']}")
                
                # Also check top-level data_payload for direct fields (fallback)
                for key, value in cert.data_payload.items():
                    if key not in ["version", "data", "signature"] and not isinstance(value, dict):
                        field_values[key] = str(value) if value is not None else ""
            
            # Step 2: Render fresh PDF with all data
            fresh_pdf_path = f"generated_certs/{cert_id}_fresh.pdf"
            try:
                pdf_utils.render_pdf_certificate(
                    pdf_template_path, 
                    field_values, 
                    fresh_pdf_path,
                    metadata={"cert_id": cert_id}
                )
                print(f"DEBUG: Fresh PDF rendered with all data: {fresh_pdf_path}")
            except Exception as e:
                print(f"ERROR: Fresh PDF render failed: {e}")
                # Fallback to existing PDF
                fresh_pdf_path = cert.rendered_pdf_path or pdf_template_path
            
            # Step 3: Apply signatures to the fresh PDF
            try:
                pdf_utils.apply_signatures_to_pdf(
                    pdf_path=fresh_pdf_path,
                    signature_img_path=sig_path,
                    stamp_img_path=stamp_path,
                    template_path=pdf_template_path,
                    output_path=signed_pdf_path,
                    signer_info={
                        "name": signer_name,
                        "role": signer_role
                    },
                    metadata={"cert_id": cert_id}
                )
                cert.rendered_pdf_path = signed_pdf_path
                print(f"DEBUG: Successfully signed PDF certificate {cert_id}")
            except Exception as e:
                print(f"ERROR: Signing failed for {cert_id}: {e}")
                continue
        elif cert.template_type == "html":
            # Use HTML template for signing
            print(f"DEBUG: Signing HTML certificate {cert_id} using HTML template")
            
            verify_url = f"{FRONTEND_URL}/verify?id={cert.id}"
            qr_b64 = generate_qr_base64(verify_url)
            
            try:
                from jinja2 import Template
                tmpl = Template(builtin_templates.get_builtin_template_html(cert.cert_type or "certificate"))

                # Encode signature/stamp images as base64 so they render inline
                def _img_b64(path):
                    if path and os.path.exists(path):
                        try:
                            with open(path, "rb") as imgf:
                                return base64.b64encode(imgf.read()).decode("utf-8")
                        except OSError:
                            return ""
                    return ""

                render_ctx = {
                    "student_name": cert.student_name,
                    "course_name": cert.course_name,
                    "issued_at": cert.issued_at.strftime("%Y-%m-%d") if cert.issued_at else datetime.datetime.now().strftime("%Y-%m-%d"),
                    "cert_id": cert.id,
                    "signature": cert.signature[:30] + "..." if cert.signature else "N/A",
                    "qr_code": qr_b64,
                    "signer_name": signer_name,
                    "signer_role": signer_role,
                    "digital_signature": _img_b64(sig_path),
                    "stamp": _img_b64(stamp_path),
                }

                # Merge any custom fields stored on the certificate's data_payload
                if cert.data_payload and isinstance(cert.data_payload, dict):
                    oa_data = cert.data_payload.get("data", {})
                    if isinstance(oa_data, dict):
                        for key, salt_value_obj in oa_data.items():
                            if isinstance(salt_value_obj, dict) and "value" in salt_value_obj:
                                if key not in ["id", "type", "name", "issuedOn", "recipient.name", "recipient.studentId", "issuers"] and key not in render_ctx:
                                    render_ctx[key] = str(salt_value_obj["value"]) if salt_value_obj["value"] is not None else ""

                html_content = tmpl.render(**render_ctx)
                
                # Convert HTML to PDF
                result = BytesIO()
                pisa.pisaDocument(BytesIO(html_content.encode("utf-8")), result)
                with open(signed_pdf_path, "wb") as f:
                    f.write(result.getvalue())
                cert.rendered_pdf_path = signed_pdf_path
                print(f"DEBUG: Successfully signed HTML certificate {cert_id}")
            except Exception as e:
                print(f"ERROR: HTML signing failed for {cert_id}: {e}")
                continue
        else:
            # Fallback: try to use PDF template if available
            if has_pdf_template:
                print(f"DEBUG: Fallback - using PDF template for certificate {cert_id}")
                try:
                    # Generate PDF from template with certificate data
                    verify_url = f"{FRONTEND_URL}/verify?id={cert.id}"
                    qr_b64 = generate_qr_base64(verify_url)
                    issued_at_str = cert.issued_at.strftime("%Y-%m-%d") if cert.issued_at else datetime.datetime.now().strftime("%Y-%m-%d")
                    
                    # Build field values for PDF template
                    field_values = {
                        "student_name": cert.student_name,
                        "name": cert.student_name,
                        "recipient": cert.student_name,
                        "recipient_name": cert.student_name, # PDF template field
                        "course_name": cert.course_name,
                        "course": cert.course_name,
                        "subject": cert.course_name,
                        "issued_at": issued_at_str,
                        "date": issued_at_str,
                        "cert_id": cert.id,
                        "id": cert.id,
                        "id8": cert.id[:8],
                        "signature": cert.signature[:20] + "..." if cert.signature else "N/A",
                        "qr_code": qr_b64,
                    }
                    
                    # Extract additional fields from certificate's data_payload
                    if cert.data_payload and isinstance(cert.data_payload, dict):
                        # Handle OpenAttestation document structure
                        oa_data = cert.data_payload.get("data", {})
                        if isinstance(oa_data, dict):
                            # Extract values from OA salt/value structure
                            for key, salt_value_obj in oa_data.items():
                                if isinstance(salt_value_obj, dict) and "value" in salt_value_obj:
                                    # Skip system fields and nested objects
                                    if key not in ["id", "type", "name", "issuedOn", "recipient.name", "recipient.studentId", "issuers"]:
                                        field_values[key] = str(salt_value_obj["value"]) if salt_value_obj["value"] is not None else ""
                            
                            # Specifically ensure dept_head is mapped from OA structure
                            dept_head_obj = (oa_data.get("dept_head") or 
                                           oa_data.get("department_head") or 
                                           oa_data.get("head_of_department") or 
                                           oa_data.get("hod"))
                            if dept_head_obj and isinstance(dept_head_obj, dict) and "value" in dept_head_obj:
                                field_values["dept_head"] = str(dept_head_obj["value"])
                        
                        # Also check top-level data_payload for direct fields (fallback)
                        for key, value in cert.data_payload.items():
                            if key not in ["version", "data", "signature"] and not isinstance(value, dict):
                                field_values[key] = str(value) if value is not None else ""
                    
                    # First render the base PDF with data
                    base_pdf_path = f"generated_certs/{cert_id}_base.pdf"
                    pdf_utils.render_pdf_certificate(
                        pdf_template_path, 
                        field_values, 
                        base_pdf_path,
                        metadata={"cert_id": cert_id}
                    )
                    
                    # Then apply signatures
                    pdf_utils.apply_signatures_to_pdf(
                        pdf_path=base_pdf_path,
                        signature_img_path=sig_path,
                        stamp_img_path=stamp_path,
                        template_path=pdf_template_path,
                        output_path=signed_pdf_path,
                        signer_info={
                            "name": signer_name,
                            "role": signer_role
                        },
                        metadata={"cert_id": cert_id}
                    )
                    cert.rendered_pdf_path = signed_pdf_path
                    print(f"DEBUG: Successfully signed certificate {cert_id} using fallback PDF method")
                except Exception as e:
                    print(f"ERROR: Fallback signing failed for {cert_id}: {e}")
                    continue
            else:
                print(f"ERROR: No suitable template found for certificate {cert_id}")
                continue

        # Update signing metadata
        existing_sigs = cert.digital_signatures or []
        existing_sigs.append({
            "signer_name": signer_name,
            "signer_role": signer_role,
            "applied_at": now_iso
        })
        cert.digital_signatures = existing_sigs
        cert.signing_status = "signed"
        
        # UNIQUE PIN GENERATION ON SIGNING
        if not cert.claim_pin:
            cert.claim_pin = "".join([str(random.randint(0, 9)) for _ in range(6)])
        
        # ═══════════════════════════════════════════════════════════════════
        # ADD WPS-STYLE INTERACTIVE VERIFICATION RIBBON
        # ═══════════════════════════════════════════════════════════════════
        try:
            if cert.rendered_pdf_path and os.path.exists(cert.rendered_pdf_path):
                # Prepare certificate data for WPS-style ribbon
                cert_data = {
                    'id': cert.id,
                    'student_name': cert.student_name,
                    'course_name': cert.course_name,
                    'issued_at': cert.issued_at.strftime('%Y-%m-%d') if cert.issued_at else datetime.datetime.now().strftime('%Y-%m-%d'),
                    'organization': cert.organization or 'EduCerts',
                    'content_hash': cert.content_hash,
                    'signature': cert.signature
                }
                
                # Add verification badge to PDF with gap above certificate
                wps_enhanced_pdf_path = f"generated_certs/{cert_id}_verified.pdf"
                try:
                    add_final_ribbon(
                        cert.rendered_pdf_path,
                        wps_enhanced_pdf_path,
                        cert_data
                    )
                    # Update certificate to point to verified PDF
                    cert.rendered_pdf_path = wps_enhanced_pdf_path
                    
                    # IMPORTANT: Update the content hash to match the final PDF with ribbon
                    try:
                        final_pdf_hash = pdf_hash_utils.compute_pdf_content_hash(wps_enhanced_pdf_path)
                        cert.content_hash = final_pdf_hash
                        print(f"✅ Updated content hash for final PDF with ribbon: {final_pdf_hash[:8]}...")
                    except Exception as hash_err:
                        print(f"⚠️  Failed to update content hash for final PDF: {hash_err}")
                    
                    print(f"✅ Added verification badge to {cert_id}")
                except Exception as ribbon_err:
                    print(f"⚠️  Failed to add verification badge to {cert_id}: {ribbon_err}")
                    # Continue without badge - signing still successful
                
        except Exception as e:
            print(f"⚠️  Error processing certificate {cert_id}: {e}")
            # Continue - signing still successful
            
        signed_certs.append({"id": cert_id, "student_name": cert.student_name, "pin": cert.claim_pin})

    db.commit()
    return {
        "message": f"{len(signed_certs)} certificates signed",
        "signed": signed_certs
    }


@app.post("/api/sign/apply-batch/{batch_id}")
async def apply_signatures_to_batch(
    batch_id: str,
    body: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin)
):
    """Apply digital signature to all certs in a batch."""
    certs = db.query(models.Certificate).filter(
        models.Certificate.batch_id == batch_id
    ).all()
    cert_ids = [c.id for c in certs]
    body["cert_ids"] = cert_ids
    # Delegate to apply endpoint logic
    return await apply_digital_signatures(body, db, current_user)


@app.get("/api/certificates/unsigned")
def get_unsigned_certificates(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin)
):
    """Returns all certificates that have not yet been digitally signed."""
    certs = db.query(models.Certificate).filter(
        models.Certificate.signing_status == "unsigned",
        models.Certificate.revoked == False
    ).order_by(models.Certificate.issued_at.desc()).all()
    return [
        {
            "id": c.id,
            "student_name": c.student_name,
            "course_name": c.course_name,
            "cert_type": c.cert_type,
            "issued_at": c.issued_at.isoformat() if c.issued_at else None,
            "organization": c.organization,
            "signing_status": c.signing_status,
            "template_type": c.template_type
        }
        for c in certs
    ]


@app.get("/api/certificates/{cert_id}")
def get_certificate(cert_id: str, db: Session = Depends(get_db)):
    """Returns details for a specific certificate."""
    c = db.query(models.Certificate).filter(models.Certificate.id == cert_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Certificate not found")
    return {
        "id": c.id,
        "student_name": c.student_name,
        "course_name": c.course_name,
        "cert_type": c.cert_type,
        "issued_at": c.issued_at.isoformat() if c.issued_at else None,
        "organization": c.organization,
        "signing_status": c.signing_status,
        "template_type": c.template_type
    }


@app.get("/api/sign/records")
def get_signature_records(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin)
):
    """Return all uploaded signature/stamp records."""
    records = db.query(models.DigitalSignatureRecord).order_by(
        models.DigitalSignatureRecord.uploaded_at.desc()
    ).all()
    return [
        {
            "id": r.id,
            "signer_name": r.signer_name,
            "signer_role": r.signer_role,
            "has_signature": r.signature_path is not None and os.path.exists(r.signature_path or ""),
            "has_stamp": r.stamp_path is not None and os.path.exists(r.stamp_path or ""),
            "uploaded_at": r.uploaded_at.isoformat() if r.uploaded_at else None
        }
        for r in records
    ]


@app.post("/api/certificates/{cert_id}/add-ribbon")
async def add_ribbon_to_certificate(
    cert_id: str,
    ribbon_config: dict = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin)
):
    """
    Add interactive verification ribbon to an existing certificate.
    
    Body: {
        "theme": "classic_blue" | "emerald_green" | "ruby_red" | "gold_premium" | "slate_professional",
        "position": "top_left" | "top_right" | "top_center" | "bottom_left" | "bottom_right" | "bottom_center",
        "custom_colors": {
            "primary": "#2563eb",
            "accent": "#d4af37"
        }
    }
    """
    cert = db.query(models.Certificate).filter(models.Certificate.id == cert_id).first()
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")
    
    if not cert.rendered_pdf_path or not os.path.exists(cert.rendered_pdf_path):
        raise HTTPException(status_code=400, detail="Certificate PDF not found")
    
    try:
        # DISABLED: Remove all ribbon functionality for clean PDFs
        # The WPS ribbon will be added only during the signing step
        
        # Just return the original PDF path without modifications
        ribbon_output_path = cert.rendered_pdf_path
        ribbon_success = True
        
        if ribbon_success:
            # Update certificate path
            cert.rendered_pdf_path = ribbon_output_path
            db.commit()
            
            return {
                "message": "Interactive ribbon added successfully",
                "certificate_id": cert_id,
                "ribbon_path": ribbon_output_path,
                "theme": theme_name,
                "position": position_name
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to add ribbon to certificate")
            
    except Exception as e:
        print(f"ERROR: Failed to add ribbon to certificate {cert_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Ribbon addition failed: {str(e)}")


@app.get("/api/certificates/{cert_id}/ribbon-preview")
async def preview_certificate_ribbon(
    cert_id: str,
    theme: str = "classic_blue",
    position: str = "top_left",
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_user)
):
    """
    Generate a preview of how the ribbon would look on a certificate.
    Returns ribbon configuration and styling information.
    """
    cert = db.query(models.Certificate).filter(models.Certificate.id == cert_id).first()
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")
    
    # Map theme and position
    theme_map = {
        "classic_blue": RibbonTheme.CLASSIC_BLUE,
        "emerald_green": RibbonTheme.EMERALD_GREEN,
        "ruby_red": RibbonTheme.RUBY_RED,
        "gold_premium": RibbonTheme.GOLD_PREMIUM,
        "slate_professional": RibbonTheme.SLATE_PROFESSIONAL
    }
    
    position_map = {
        "top_left": RibbonPosition.TOP_LEFT,
        "top_right": RibbonPosition.TOP_RIGHT,
        "top_center": RibbonPosition.TOP_CENTER,
        "bottom_left": RibbonPosition.BOTTOM_LEFT,
        "bottom_right": RibbonPosition.BOTTOM_RIGHT,
        "bottom_center": RibbonPosition.BOTTOM_CENTER
    }
    
    ribbon_theme = theme_map.get(theme, RibbonTheme.CLASSIC_BLUE)
    ribbon_position = position_map.get(position, RibbonPosition.TOP_LEFT)
    
    # Create ribbon style
    ribbon_style = RibbonStyleManager.create_theme_style(ribbon_theme, ribbon_position)
    
    # Create verification result for preview
    verification_request = schemas.VerificationRequest(certificate_id=cert_id)
    verification_result = verify_certificate(verification_request, db)
    
    # Create verification metadata
    metadata = verification_metadata.create_verification_metadata_from_api_result(cert, verification_result)
    
    return {
        "certificate_id": cert_id,
        "ribbon_style": ribbon_style.to_dict(),
        "verification_status": metadata.get_verification_status_text(),
        "is_fully_verified": metadata.is_fully_verified(),
        "preview_data": {
            "theme": theme,
            "position": position,
            "colors": ribbon_style.colors.to_dict(),
            "dimensions": ribbon_style.dimensions.to_dict()
        }
    }


@app.post("/api/certificates/batch-add-ribbons")
async def batch_add_ribbons_to_certificates(
    body: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin)
):
    """
    Add ribbons to multiple certificates in batch.
    
    Body: {
        "cert_ids": ["cert1", "cert2", ...],
        "ribbon_config": {
            "theme": "classic_blue",
            "position": "top_left"
        }
    }
    """
    cert_ids = body.get("cert_ids", [])
    ribbon_config = body.get("ribbon_config", {})
    
    if not cert_ids:
        raise HTTPException(status_code=400, detail="No certificate IDs provided")
    
    results = []
    
    for cert_id in cert_ids:
        try:
            # Use the single certificate ribbon endpoint
            result = await add_ribbon_to_certificate(cert_id, ribbon_config, db, current_user)
            results.append({
                "certificate_id": cert_id,
                "status": "success",
                "message": result["message"]
            })
        except Exception as e:
            results.append({
                "certificate_id": cert_id,
                "status": "error",
                "message": str(e)
            })
    
    successful = len([r for r in results if r["status"] == "success"])
    
    return {
        "message": f"Processed {len(cert_ids)} certificates, {successful} successful",
        "results": results,
        "total_processed": len(cert_ids),
        "successful": successful,
        "failed": len(cert_ids) - successful
    }


@app.get("/api/import")
async def import_data(file: UploadFile = File(...)):
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files are allowed")
    content = await file.read()
    csv_reader = csv.DictReader(io.StringIO(content.decode('utf-8')))
    data = [row for row in csv_reader if "student_name" in row and "course_name" in row]
    return {"message": "Data imported successfully", "count": len(data), "data": data}

@app.get("/api/preview-signature/{cert_id}/{sig_record_id}")
async def preview_signature(
    cert_id: str,
    sig_record_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin)
):
    """
    Generate a temporary preview of the certificate with the signature/stamp
    overlaid. Returns a PNG image of the first page without saving permanently.
    """
    import tempfile, fitz  # PyMuPDF
    try:
        cert = db.query(models.Certificate).filter(models.Certificate.id == cert_id).first()
        if not cert:
            raise HTTPException(status_code=404, detail="Certificate not found")

        sig_record = db.query(models.DigitalSignatureRecord).filter(
            models.DigitalSignatureRecord.id == sig_record_id
        ).first()
        if not sig_record:
            raise HTTPException(status_code=404, detail="Signature record not found")

        sig_path = sig_record.signature_path
        stamp_path = sig_record.stamp_path
        pdf_template_path = "user_templates/template.pdf"

        os.makedirs("generated_certs", exist_ok=True)

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False, dir="generated_certs") as tmp:
            tmp_path = tmp.name

        if cert.template_type == "pdf" and os.path.exists(pdf_template_path):
            # Start from the already-rendered base PDF or re-render from template
            base_path = cert.rendered_pdf_path if (cert.rendered_pdf_path and os.path.exists(cert.rendered_pdf_path)) else None

            if base_path:
                pdf_utils.apply_signatures_to_pdf(
                    pdf_path=base_path,
                    signature_img_path=sig_path,
                    stamp_img_path=stamp_path,
                    template_path=pdf_template_path,
                    output_path=tmp_path,
                    signer_info={
                        "name": sig_record.signer_name,
                        "role": sig_record.signer_role
                    },
                    metadata={"cert_id": cert.id}
                )
            else:
                verify_url = f"{FRONTEND_URL}/verify?id={cert.id}"
                qr_b64 = generate_qr_base64(verify_url)
                field_values = {
                    "student_name": cert.student_name,
                    "course_name": cert.course_name,
                    "issued_at": cert.issued_at.strftime("%Y-%m-%d") if cert.issued_at else "",
                    "cert_id": cert.id,
                    "signature": (cert.signature or "")[:20] + "...",
                    "qr_code": qr_b64,
                }
                intermediate = tmp_path + "_base.pdf"
                pdf_utils.render_pdf_certificate(
                    pdf_template_path, field_values, intermediate,
                    signature_img_path=sig_path,
                    stamp_img_path=stamp_path,
                    metadata={"cert_id": cert.id}
                )
                import shutil
                shutil.move(intermediate, tmp_path)
        else:
            verify_url = f"{FRONTEND_URL}/verify?id={cert.id}"
            qr_b64 = generate_qr_base64(verify_url)
            from jinja2 import Template
            tmpl = Template(builtin_templates.get_builtin_template_html(cert.cert_type or "certificate"))
            
            render_ctx = {
                "student_name": cert.student_name,
                "course_name": cert.course_name,
                "issued_at": cert.issued_at.strftime("%Y-%m-%d") if cert.issued_at else "",
                "cert_id": cert.id,
                "signature": (cert.signature or "")[:30] + "...",
                "qr_code": qr_b64,
            }
            html_content = tmpl.render(**render_ctx)
            result = BytesIO()
            pisa.pisaDocument(BytesIO(html_content.encode("utf-8")), result)
            with open(tmp_path, "wb") as f:
                f.write(result.getvalue())

        # Convert first page to PNG and return it
        doc = fitz.open(tmp_path)
        page = doc[0]
        mat = fitz.Matrix(2.0, 2.0)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        png_bytes = pix.tobytes("png")
        doc.close()

        return Response(content=png_bytes, media_type="image/png")

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Preview generation failed: {e}")
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass


@app.get("/api/view/{cert_id}")
def view_certificate(cert_id: str, db: Session = Depends(get_db)):
    """View certificate inline (not download)"""
    cert = db.query(models.Certificate).filter(models.Certificate.id == cert_id).first()
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")

    # Check if rendered PDF exists.
    # For unsigned HTML certificates, re-render so template improvements apply.
    should_rerender_html_unsigned = cert.template_type == "html" and cert.signing_status != "signed"
    if cert.rendered_pdf_path and os.path.exists(cert.rendered_pdf_path) and not should_rerender_html_unsigned:
        return FileResponse(
            path=cert.rendered_pdf_path,
            media_type="application/pdf",
            headers={"Content-Disposition": f"inline; filename=cert_{cert.id}.pdf"}
        )
    
    # If no rendered PDF, try to generate it on the fly
    pdf_template_path = "user_templates/template.pdf"
    if cert.template_type == "pdf" and os.path.exists(pdf_template_path):
        verify_url = f"{FRONTEND_URL}/verify?id={cert.id}"
        qr_b64 = generate_qr_base64(verify_url)
        field_values = {
            "student_name": cert.student_name,
            "name": cert.student_name,
            "course_name": cert.course_name,
            "course": cert.course_name,
            "issued_at": cert.issued_at.strftime("%Y-%m-%d"),
            "date": cert.issued_at.strftime("%Y-%m-%d"),
            "cert_id": cert.id,
            "signature": cert.signature[:20] + "..." if cert.signature else "Unsigned",
            "qr_code": qr_b64,
        }
        
        os.makedirs("generated_certs", exist_ok=True)
        out_path = f"generated_certs/{cert.id}_view.pdf"
        
        try:
            pdf_utils.render_pdf_certificate(
                pdf_template_path, 
                field_values, 
                out_path,
                metadata={"cert_id": cert.id}
            )
            
            return FileResponse(
                path=out_path,
                media_type="application/pdf",
                headers={"Content-Disposition": f"inline; filename=cert_{cert.id}.pdf"}
            )
        except Exception as e:
            print(f"Error rendering PDF: {e}")
            raise HTTPException(status_code=500, detail=f"Error rendering PDF: {str(e)}")

    # HTML fallback for type-based built-in templates
    if cert.template_type != "pdf":
        verify_url = f"{FRONTEND_URL}/verify?id={cert.id}"
        qr_base64 = generate_qr_base64(verify_url)

        from jinja2 import Template
        template = Template(builtin_templates.get_builtin_template_html(cert.cert_type or "certificate"))

        render_ctx = {
            "student_name": cert.student_name,
            "course_name": cert.course_name,
            "issued_at": cert.issued_at.strftime("%Y-%m-%d") if cert.issued_at else "",
            "cert_id": cert.id,
            "signature": cert.signature[:30] + "..." if cert.signature else "Unsigned",
            "qr_code": qr_base64,
        }

        payload_data = cert.data_payload or {}
        extra_fields = {}
        for k, v in payload_data.items():
            if k not in ("signature", "data", "schema") and isinstance(v, (str, int, float)):
                extra_fields[k] = v
        oa_data = payload_data.get("data", {})
        if isinstance(oa_data, dict):
            for k, v in oa_data.items():
                if isinstance(v, dict) and "value" in v:
                    extra_fields[k] = v["value"]
                elif isinstance(v, (str, int, float)):
                    extra_fields[k] = v

        html_content = template.render(**{**extra_fields, **render_ctx})
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html_content.encode("utf-8")), result)
        if pdf.err:
            raise HTTPException(status_code=500, detail="Error generating certificate PDF")

        return Response(
            content=result.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": f"inline; filename=cert_{cert.id}.pdf"}
        )

    raise HTTPException(status_code=404, detail="Certificate PDF not found.")


@app.get("/api/download/{cert_id}")
def download_certificate(cert_id: str, db: Session = Depends(get_db)):
    cert = db.query(models.Certificate).filter(models.Certificate.id == cert_id).first()
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")

    # ── If a rendered/signed PDF already exists, serve it directly. ──
    # Re-render unsigned HTML certificates so new type templates are reflected.
    should_rerender_html_unsigned = cert.template_type == "html" and cert.signing_status != "signed"
    if cert.rendered_pdf_path and os.path.exists(cert.rendered_pdf_path) and not should_rerender_html_unsigned:
        return FileResponse(
            path=cert.rendered_pdf_path,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=cert_{cert.id}.pdf"}
        )

    # ── PDF template path ──
    pdf_template_path = "user_templates/template.pdf"
    if cert.template_type == "pdf" and os.path.exists(pdf_template_path):
        # Render on-the-fly from PDF template
        verify_url = f"{FRONTEND_URL}/verify?id={cert.id}"
        qr_b64 = generate_qr_base64(verify_url)
        field_values = {
            "student_name": cert.student_name,
            "name": cert.student_name,
            "course_name": cert.course_name,
            "course": cert.course_name,
            "issued_at": cert.issued_at.strftime("%Y-%m-%d"),
            "date": cert.issued_at.strftime("%Y-%m-%d"),
            "cert_id": cert.id,
            "signature": cert.signature[:20] + "...",
            "qr_code": qr_b64,
        }
        # Also overlay payload fields - ROBUST EXTRACTION
        payload_data = cert.data_payload or {}
        # Try both direct payload and OA 'data' nested payload
        candidates = [payload_data, payload_data.get("data", {})]
        
        for source in candidates:
            if not isinstance(source, dict): continue
            for k, v in source.items():
                # OA might have values like {"value": "John"} or just "John"}
                if isinstance(v, dict) and "value" in v:
                    field_values.setdefault(k, v["value"])
                elif isinstance(v, (str, int, float)):
                    field_values.setdefault(k, v)
                elif isinstance(v, dict):
                    # Check nested objects like recipient.name
                    for subk, subv in v.items():
                        if isinstance(subv, (str, int, float)):
                            field_values.setdefault(f"{k}_{subk}", subv)
                            field_values.setdefault(subk, subv)

        field_values["id"] = cert.id
        field_values["id8"] = cert.id[:8]

        print(f"DEBUG: On-the-fly field values: {list(field_values.keys())}")
        os.makedirs("generated_certs", exist_ok=True)
        out_path = f"generated_certs/{cert.id}_base.pdf"
        try:
            # We don't have the map cached globally, but we only do this once per download
            pdf_utils.render_pdf_certificate(
                pdf_template_path, 
                field_values, 
                out_path,
                metadata={"cert_id": cert.id}
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"PDF render error: {e}")

        # Save the rendered path for next time
        cert.rendered_pdf_path = out_path
        db.commit()

        return FileResponse(
            path=out_path,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=cert_{cert.id}.pdf"}
        )

    # ── Fallback: HTML template → xhtml2pdf ──
    # ONLY if it's not a PDF template. If it's a PDF template and we are here, something is wrong.
    if cert.template_type == "pdf":
        raise HTTPException(status_code=500, detail="PDF template was requested but rendering failed or template is missing.")

    verify_url = f"{FRONTEND_URL}/verify?id={cert.id}"
    qr_base64 = generate_qr_base64(verify_url)

    from jinja2 import Template
    template = Template(builtin_templates.get_builtin_template_html(cert.cert_type or "certificate"))

    render_ctx = {
        "student_name": cert.student_name,
        "course_name": cert.course_name,
        "issued_at": cert.issued_at.strftime("%Y-%m-%d"),
        "cert_id": cert.id,
        "signature": cert.signature[:30] + "...",
        "qr_code": qr_base64,
    }

    payload_data = cert.data_payload or {}
    extra_fields = {}
    for k, v in payload_data.items():
        if k not in ("signature", "data", "schema") and isinstance(v, (str, int, float)):
            extra_fields[k] = v
    oa_data = payload_data.get("data", {})
    if isinstance(oa_data, dict):
        for k, v in oa_data.items():
            if isinstance(v, dict) and "value" in v:
                extra_fields[k] = v["value"]
            elif isinstance(v, (str, int, float)):
                extra_fields[k] = v

    render_ctx = {**extra_fields, **render_ctx}
    html_content = template.render(**render_ctx)

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode("utf-8")), result)
    if pdf.err:
        raise HTTPException(status_code=500, detail="Error generating PDF")

    return Response(
        content=result.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=cert_{cert.id}.pdf"}
    )

@app.get("/api/json/{cert_id}")
def download_json_certificate(cert_id: str, db: Session = Depends(get_db)):
    cert = db.query(models.Certificate).filter(models.Certificate.id == cert_id).first()
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")
    return JSONResponse(
        content=cert.data_payload,
        headers={"Content-Disposition": f"attachment; filename=cert_{cert.id}.json"}
    )
